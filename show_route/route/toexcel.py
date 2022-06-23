from numpy.random import random
from .models import SavedRoutes
from openpyxl import Workbook
from os.path import join


def saveToXlsx(start, end, xlsx_data):
    start_lat = start[0]
    start_long = start[1]
    end_lat = end[0]
    end_long = end[1]

    filename = str(random(1)) + '.xlsx'
    path = join('routes', filename)
    workbook = Workbook()
    a = SavedRoutes()

    worksheet = workbook.active
    worksheet.title = 'Route'

    worksheet['A1'] = 'Стартовая точка'
    worksheet['B1'] = start_lat
    worksheet['C1'] = start_long
    worksheet['A2'] = 'Конечная точка'
    worksheet['B2'] = end_lat
    worksheet['C2'] = end_long

    columns = [
        'Way_ID',
        'Point_lat',
        'Point_long',
        'Street'
    ]
    row_num = 4

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for data in xlsx_data:
        row_num += 1

        row = [
            data['id'],
            data['lat'],
            data['long'],
            data['street']
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(path)

    a.start_point_lat = start_lat
    a.start_point_long = start_long
    a.end_point_lat = end_lat
    a.end_point_long = end_long
    a.route_file.name = path

    a.save()

    return True
